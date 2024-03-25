import boto3
import json
import pandas as pd
from textractcaller import call_textract, Textract_Features
from textractprettyprinter.t_pretty_print import Pretty_Print_Table_Format, Textract_Pretty_Print, get_string, get_tables_string
from trp import Document
from trp.trp2 import TDocument, TDocumentSchema
from trp.t_pipeline import order_blocks_by_geo
from IPython.display import display
import openpyxl
from openpyxl.styles import Font
import os
from dotenv import load_dotenv

load_dotenv()

s3_file = 's3://textract-testing-sts004/GMC1095c1_update1.pdf'
session = boto3.client('textract',
                      region_name='us-east-1',
                      aws_access_key_id=os.getenv('ACCESS_KEY_ID'),
                      aws_secret_access_key=os.getenv('SECRET_ACCESS_KEY'))
textract_json = call_textract(input_document=s3_file, features = [Textract_Features.TABLES,Textract_Pretty_Print.FORMS],boto3_textract_client=session)

required_data =  []

with open('/home/sts852-aadhithyar/Documents/ACA/Main/ACA_Main/Form_field1.txt', 'r') as f:
    for line in f:
        required_data.append(line.strip())

doc = Document(textract_json)
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string


# Create a new Excel workbook
workbook = openpyxl.load_workbook('Final.xlsx')

sheet = workbook.active
global temp_row, temp_col, temp
# temp is used to check whether part 3 is present or not it is used to check part 2 pages continues or not.
temp = 1
temp_row = 3
temp_col = 1
# Select the active sheet (create new sheet for each page)
for page_idx, page in enumerate(doc.pages):
    tables = page.tables
    #part_1 dictionary has all the data form the form
    part_1 = {}
    #main dictionary contains only the required data which is needed
    main = {}
    #temp_col = 1
    if len(tables) == 2:
        # temp is checked whether it is 0 or not, if it is 0 then the previous page is also part 2 contnet.
        if(temp == 0):
          temp_col = 1
        table_0 = tables[0]
        table_1 = tables[1]

        # Data stored in dictionary
        for field in page.form.fields:
            part_1[str(field.key)] = str(field.value)

        # Getting the required Data
        for i in required_data:
            main[i] = part_1.get(i, "")

        l = []  # Initialize list outside the loop to store rows
        for row in table_1.rows:
            # Initialize an empty list to store the content of cells in the row
            temp = []
            for cell in row.cells:
                # Append the content of each cell to the temporary list
                temp.append(cell.text)
            # Append the content of the row to the main list
            l.append(temp)

        # Write the keys and values from the dictionary to the Excel file
        row = temp_row
        col = temp_col
        for key, value in main.items():
            if(key == "1 Name of employee (first name, middle initial, last name)"):
              temp = value.split(" ")
              if(len(temp) == 2):
                temp.insert(1," ")
              for i in temp:
                sheet.cell(row=row, column=col).value = i
                col += 1
            else:
                sheet.cell(row=row, column=col).value = value
                col += 1


# Part II
        # Write the table data to the Excel file
        for row_idx, row_data in enumerate(l[1:len(l)], start=row):  # Exclude the first row
            for col_idx, cell_value in enumerate(row_data):
                # This is the condition which is used to check for the value is 0 dollors.
                if(len(cell_value) == 3 or (len(cell_value) > 0 and (ord(cell_value[0]) == 36 or cell_value[0] == "0")) or len(cell_value) == 0 or len(cell_value) == 6):
                    sheet.cell(row=row, column=col+1).value = cell_value
                    col += 1
        temp_row = row
        temp = 0
        temp_col = col

    elif len(tables) == 1:
# Part III
        # The temp is used to set the last page is set tat the last page is part 3.
        temp = 1
        row = temp_row -1
        col = temp_col
        table = tables[0]
        all_rows = []
        #start is used for the multiple people in the part 3 content.
        start = temp_col
        for i, rows in enumerate(table.rows):
            if i <= 1:
                continue
            all_rows.append([cell.text for cell in rows.cells])
        for row_idx, row_data in enumerate(all_rows, start=row):
            if(row_data[1] != ""):
              #this is reset again so that the next person is writtien in the same line in next row
              col = start
              for col_idx, cell_value in enumerate(row_data[1:]):
                  sheet.cell(row=row, column= col+1).value = cell_value
                  col += 1
              row += 1
            #row -= 1
        temp_col = 1
    temp_row = row +1

sheet.delete_cols(13, 2)

rows_to_delete = []
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    if all(cell.value is None for cell in row):
        rows_to_delete.append(row[0].row)

# Delete rows in reverse order to avoid index issues
for row_index in sorted(rows_to_delete, reverse=True):
    sheet.delete_rows(row_index)


workbook.save('/home/sts852-aadhithyar/Documents/ACA/Main/ACA_Main/Test.xlsx')